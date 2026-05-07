# Import necessary libraries
import streamlit as st
from mesa import Agent, Model
from mesa.time import RandomActivation
from model import NeuralNet, ChatBotModel
from pracbot import bag_of_words, tokenize
import torch
import json
import random
import pandas as pd
import datetime as dt
import uuid
import openpyxl

# Initialize the uploaded_file variable
excel_file = r"C:\Users\murar\Downloads\Book.xlsx"
uploaded_file = None

try:
    uploaded_file = openpyxl.load_workbook(excel_file)
    bookings_df = pd.read_excel(uploaded_file, engine='openpyxl')
except ValueError:
    # If the 'openpyxl' engine fails, try using the 'xlrd' engine for older Excel files
    uploaded_file = openpyxl.load_workbook(excel_file, engine='xlrd')
    bookings_df = pd.read_excel(uploaded_file, engine='xlrd')
except FileNotFoundError:
    # Create a new DataFrame if the file doesn't exist
    bookings_df = pd.DataFrame(columns=['booking_id', 'doctor', 'date', 'timing'])

# Define the UserAgent class
class UserAgent(Agent):
    def __init__(self, unique_id, model, chatbot):
        super().__init__(unique_id, model)
        self.chatbot = chatbot
        self.input_key = f"user_input_{unique_id}"  # Unique key for each agent

    def step(self):
        # User agent interacts with the chatbot
        user_input = f"User {self.unique_id}: {st.text_input(f'User {self.unique_id}:', key=self.input_key)}"
        self.chatbot.process_user_input(user_input)

# Define the ChatBotModel class
class ChatBotModel(Model):
    def __init__(self, num_agents, chatbot):
        self.num_agents = num_agents
        self.schedule = RandomActivation(self)
        self.chatbot = chatbot

        # Create user agents
        for i in range(self.num_agents):
            agent = UserAgent(i, self, chatbot)
            self.schedule.add(agent)

    def step(self):
        # Step through user agents
        self.schedule.step()

import streamlit as st
import random
import json
import torch
from model import NeuralNet
from pracbot import bag_of_words, tokenize
import pandas as pd
import datetime
import uuid
import openpyxl
import datetime as dt
import time
from twilio.rest import Client

# Load AI model and data
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
with open(r'C:\Users\murar\eclipse-workspace\Project\src\Murari\aichatbot\knowledge_base.json') as f:
    intents = json.load(f)

FILE = "data.pth"
data = torch.load(FILE)

input_size = data["input_size"]
hidden_size = data["hidden_size"]
output_size = data["output_size"]
all_words = data["all_words"]
tags = data["tags"]
model_state = data["model_state"]

model = NeuralNet(input_size, hidden_size, output_size).to(device)
model.load_state_dict(model_state)
model.eval()

bot_name = "sunny"

# Initialize bookings_df in the global scope
bookings_df = pd.DataFrame(columns=['booking_id', 'doctor', 'date', 'timing', 'patient_number', 'patient_name', 'patient_age'])

# Upload the Excel file using Streamlit
excel_file = r"C:\Users\murar\Downloads\Book.xlsx"
uploaded_file = openpyxl.load_workbook(excel_file)

if uploaded_file is not None:
    try:
        bookings_df = pd.read_excel(uploaded_file, engine='openpyxl')
    except ValueError:
        bookings_df = pd.read_excel(uploaded_file, engine='xlrd')
    except FileNotFoundError:
        bookings_df = pd.DataFrame(columns=['booking_id', 'doctor', 'date', 'timing', 'patient_number', 'patient_name', 'patient_age'])

def is_sessions_available(doctor, date, timing):
    try:
        bookings_df = pd.read_excel(excel_file, engine='openpyxl')
    except (ValueError, FileNotFoundError):
        return True  # Assuming it's available if there's an issue reading the file

    available_sessions = bookings_df[
        (bookings_df['doctor'] == doctor) &
        (bookings_df['date'] == date) &
        (bookings_df['timing'] == timing)
    ]

    return available_sessions.empty

def save_booking_details(booking_id, doctor, date, timing, patient_number, patient_name, patient_age):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    row = 2
    while sheet.cell(row=row, column=1).value is not None:
        row += 1

    sheet.cell(row=row, column=1, value=booking_id)
    sheet.cell(row=row, column=2, value=doctor)
    sheet.cell(row=row, column=3, value=date)
    sheet.cell(row=row, column=4, value=timing)
    sheet.cell(row=row, column=5, value=patient_number)
    sheet.cell(row=row, column=6, value=patient_name)
    sheet.cell(row=row, column=7, value=patient_age)

    wb.save(excel_file)
    print(f"Appointment details for {doctor} saved successfully.")

def bookappointment():
    available_doctors = ["Dr. Ramesh", "Dr. Rajeswari", "Dr. Mani Sharma"]
    selected_doctor = st.selectbox("Select a doctor:", available_doctors)

    today = dt.date.today()
    next_30_days = today + dt.timedelta(days=30)
    selected_date = st.date_input("Select the date:", min_value=today, max_value=next_30_days)

    if selected_date.weekday() in [5, 6]:
        st.text("No appointments available on Saturdays and Sundays.")
        return

    available_timings = ["10:00 - 11:00", "11:00 - 12:00", "14:00 - 15:00", "15:00 - 16:00", "19:00 - 20:00"]
    selected_timing = st.selectbox("Select the timing:", available_timings)
    patient_number = st.text_input("Enter your mobile number:")
    patient_name = st.text_input("Enter your name:")
    patient_age = st.text_input("Enter your age:")

    if st.button("Confirm Booking"):
        if not is_sessions_available(selected_doctor, selected_date, selected_timing):
            st.text(f"The session for {selected_doctor} on {selected_date} at {selected_timing} is booked. Please choose another timing or doctor.")
            return

        booking_id = str(uuid.uuid4())
        
        save_booking_details(booking_id, selected_doctor, selected_date, selected_timing, patient_number, patient_name, patient_age)

        # Send SMS notification
        send_sms_notification(patient_name,booking_id,selected_doctor, selected_date, selected_timing, patient_number,patient_age)

        st.text(f"Booking confirmed! Your booking ID is {booking_id}")
        st.text("Thank you for using our service.")

        # Clear the booking interface after 20 seconds
        
        time.sleep(20)
        st.empty()
        st.text("The booking interface will be  been cleared.")

def send_sms_notification(patient_name,booking_id,doctor, date, timing, patient_number,patient_age):
        
    account_sid = 'AC4d7aa5360cec05a1189df847dd0fe6a2'
    auth_token = '9743612ed9e2e89bd70f0acbc1eb835f'
    client = Client(account_sid, auth_token)

    message_body = f"Appointment confirmed with {doctor} on {date} at {timing}. \n Patient details: \n Name: {patient_name} \n Age: {patient_age} \n  booking id:{booking_id} Thank you! \n  -Arogya Hospital"
    message = client.messages.create(
        body=message_body,
        from_='+13345395416',
        to=patient_number
    )

    print(f"SMS sent to {patient_number}. SID: {message.sid}")

def search_medicine():
    st.text("Sure, let me help you with that.")
    medicine_name = st.text_input("Enter the name of the medicine:")

    medicine_info = {
    'paracetamol': {
        'description': 'Paracetamol is a common pain reliever.',
        'usage': 'Used for relieving mild to moderate pain and reducing fever.',
        'side_effects': 'May cause allergic reactions in some individuals.',
        'precautions': 'Avoid alcohol while taking paracetamol.'
    },
    'aspirin': {
        'description': 'Aspirin is often used to reduce pain and inflammation.',
        'usage': 'Used for relieving pain, reducing inflammation, and preventing blood clots.',
        'side_effects': 'May cause stomach bleeding and allergic reactions.',
        'precautions': 'Consult a doctor before use, especially for individuals with bleeding disorders.'
    },
    'ibuprofen': {
        'description': 'Ibuprofen is a nonsteroidal anti-inflammatory drug (NSAID).',
        'usage': 'Used for relieving pain, reducing inflammation, and lowering fever.',
        'side_effects': 'May cause stomach upset, heartburn, and allergic reactions.',
        'precautions': 'Take with food or milk to reduce the risk of stomach upset.'
    },
    'cetirizine': {
        'description': 'Cetirizine is an antihistamine used to relieve allergy symptoms.',
        'usage': 'Used for treating allergic rhinitis (hay fever) and hives.',
        'side_effects': 'May cause drowsiness in some individuals.',
        'precautions': 'Avoid activities requiring mental alertness until the effects are known.'
    },
    'omeprazole': {
        'description': 'Omeprazole is a proton pump inhibitor (PPI).',
        'usage': 'Used for reducing stomach acid production and treating acid-related conditions.',
        'side_effects': 'May cause headache, diarrhea, and stomach pain.',
        'precautions': 'Take it at least 30 minutes before a meal.'
    },
    'amoxicillin': {
        'description': 'Amoxicillin is an antibiotic.',
        'usage': 'Used for treating bacterial infections.',
        'side_effects': 'May cause nausea, diarrhea, and allergic reactions.',
        'precautions': 'Complete the full course of antibiotics as prescribed.'
    },
    'simvastatin': {
        'description': 'Simvastatin is a statin medication.',
        'usage': 'Used for lowering cholesterol levels.',
        'side_effects': 'May cause muscle pain and liver problems.',
        'precautions': 'Avoid grapefruit and grapefruit juice while taking simvastatin.'
    },
    'levothyroxine': {
        'description': 'Levothyroxine is a thyroid hormone replacement.',
        'usage': 'Used for treating hypothyroidism (underactive thyroid).',
        'side_effects': 'May cause hair loss and weight changes.',
        'precautions': 'Take on an empty stomach, at least 30 minutes before eating.'
    },
    'metformin': {
        'description': 'Metformin is an oral diabetes medicine.',
        'usage': 'Used for treating type 2 diabetes.',
        'side_effects': 'May cause stomach upset and diarrhea.',
        'precautions': 'Take with meals to reduce gastrointestinal effects.'
    },
    'losartan': {
        'description': 'Losartan is an angiotensin II receptor blocker (ARB).',
        'usage': 'Used for treating high blood pressure.',
        'side_effects': 'May cause dizziness and changes in kidney function.',
        'precautions': 'Avoid potassium supplements while taking losartan.'
    },
    'ondansetron': {
        'description': 'Ondansetron is an antiemetic.',
        'usage': 'Used for preventing nausea and vomiting.',
        'side_effects': 'May cause headache and constipation.',
        'precautions': 'Inform your doctor of any liver problems.'
    },
    'diazepam': {
        'description': 'Diazepam is a benzodiazepine.',
        'usage': 'Used for treating anxiety and muscle spasms.',
        'side_effects': 'May cause drowsiness and dizziness.',
        'precautions': 'Avoid alcohol and activities requiring mental alertness.'
    },
    'ranitidine': {
        'description': 'Ranitidine is an H2 blocker.',
        'usage': 'Used for reducing stomach acid production.',
        'side_effects': 'May cause headache and stomach upset.',
        'precautions': 'Take as directed, especially before meals.'
    },
    'sertraline': {
        'description': 'Sertraline is an antidepressant.',
        'usage': 'Used for treating depression and anxiety disorders.',
        'side_effects': 'May cause insomnia and sexual dysfunction.',
        'precautions': 'Follow the prescribed dosage and schedule.'
    },
    'furosemide': {
        'description': 'Furosemide is a loop diuretic.',
        'usage': 'Used for treating edema and high blood pressure.',
        'side_effects': 'May cause dehydration and electrolyte imbalances.',
        'precautions': 'Monitor potassium levels during treatment.'
    },
    'clindamycin': {
        'description': 'Clindamycin is an antibiotic.',
        'usage': 'Used for treating bacterial infections.',
        'side_effects': 'May cause diarrhea and skin rash.',
        'precautions': 'Inform your doctor of any gastrointestinal issues.'
    },
    'loratadine': {
        'description': 'Loratadine is an antihistamine.',
        'usage': 'Used for relieving allergy symptoms.',
        'side_effects': 'Generally well-tolerated with minimal side effects.',
        'precautions': 'May be taken with or without food.'
    },
    'piroxicam': {
        'description': 'Piroxicam is a nonsteroidal anti-inflammatory drug (NSAID).',
        'usage': 'Used for relieving pain and inflammation.',
        'side_effects': 'May cause stomach upset and increased blood pressure.',
        'precautions': 'Take with food to reduce gastrointestinal effects.'
    },
    'gabapentin': {
        'description': 'Gabapentin is an anticonvulsant.',
        'usage': 'Used for treating neuropathic pain and seizures.',
        'side_effects': 'May cause dizziness and drowsiness.',
        'precautions': 'Follow the prescribed dosage and schedule.'
    },
    'metoprolol': {
        'description': 'Metoprolol is a beta-blocker.',
        'usage': 'Used for treating high blood pressure and angina.',
        'side_effects': 'May cause fatigue and slow heart rate.',
        'precautions': 'Do not abruptly stop taking metoprolol.'
    },
    'thyroxine': {
        'description': 'Thyroxine is a thyroid hormone.',
        'usage': 'Used for treating hypothyroidism.',
        'side_effects': 'Generally well-tolerated with proper dosage.',
        'precautions': 'Regular monitoring of thyroid function is necessary.'
    },
    # Add more medicines as needed
}


    if medicine_name.lower() in medicine_info:
        st.text(f"Here is some information about {medicine_name}: \n {medicine_info[medicine_name.lower()]}")
    else:
        st.text(f"Sorry, I couldn't find information about {medicine_name}.")

def get_appointment_details():
    st.text("Sure, please enter your appointment ID:")
    appointment_id = st.text_input("Appointment ID:")

    appointment_details = bookings_df[bookings_df['booking_id'] == appointment_id]

    if not appointment_details.empty:
        st.text(f"Appointment Details for ID {appointment_id}:")
        st.table(appointment_details)
    else:
        st.text(f"Appointment ID {appointment_id} not found.")

st.title("AI Chatbot")
st.write("Type 'quit' to exit the chat.")

user_input = st.text_input("You:")

if user_input:
    if user_input.lower() == "quit":
        st.text(f'{bot_name}: Goodbye!')
    else:
        sentence = tokenize(user_input)
        x = bag_of_words(sentence, all_words)
        x = x.reshape(1, x.shape[0])
        x = torch.from_numpy(x)

        output = model(x)
        _, predicted = torch.max(output, dim=1)
        tag = tags[predicted.item()]

        probs = torch.softmax(output, dim=1)
        prob = probs[0][predicted.item()]
        if prob.item() > 0.75:
            response = ""
            for intent in intents["intents"]:
                if tag == intent["tag"]:
                    if tag == "book appointment":
                        response = "Please enter the details"
                        st.write(f'{bot_name}: {response}')
                        bookappointment()
                    elif tag == "appointment details":
                        response = "Please enter the appointment id: "
                        st.write(f'{bot_name}: {response}')
                        get_appointment_details()
                    elif tag == "search medicine":
                        response = "Sure, let me find information about that."
                        st.write(f'{bot_name}: {response}')
                        search_medicine()
                    else:
                        response = random.choice(intent["responses"])
                        st.write(f'{bot_name}: {response}')
        else:
            st.text(f'{bot_name}: I do not understand')
